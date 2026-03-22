"""PerfDog 核心解析冒烟测试（生成最小 xlsx）。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from toolkit.core.perfdog import compare_reports, load_and_analyze
from toolkit.core.perfdog.errors import PerfDogParseError
from toolkit.core.perfdog.report_types import (
    AnalysisReport,
    Finding,
    FindingCategory,
    FindingSeverity,
    Recommendation,
    SessionSummary,
)


def _write_minimal_perfdog_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "all"
    ws.append(["PerfDog session", "com.example.game"])
    ws.append(["Device", "TestDevice"])
    ws.append(["Data_v4"])
    ws.append(["Time(ms)", "FPS", "Smooth"])
    for i in range(50):
        t = float(i * 1000)
        fps = 58.0 if 10 <= i <= 20 else 59.5
        ws.append([t, fps, 0.95])
    wb.save(path)


def test_load_and_analyze_minimal_xlsx(tmp_path: Path) -> None:
    p = tmp_path / "min.xlsx"
    _write_minimal_perfdog_xlsx(p)
    report = load_and_analyze(str(p))
    assert report.session.target_fps_hint is not None
    assert report.summary_metrics.get("采样点数") == 50
    assert len(report.findings) >= 1
    assert len(report.recommendations) >= 1


def test_reject_non_excel(tmp_path: Path) -> None:
    bad = tmp_path / "a.txt"
    bad.write_text("not excel", encoding="utf-8")
    with pytest.raises(PerfDogParseError):
        load_and_analyze(str(bad))


def _write_variants_xlsx(path: Path, *, sheet_title: str, label: str, gap_rows: int) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_title
    ws.append(["meta"])
    ws.append([label])
    for _ in range(gap_rows):
        ws.append([])
    ws.append(["Time(ms)", "SmallJank", "FPS"])
    for i in range(5):
        ws.append([float(i * 1000), 0, 59.0])
    wb.save(path)


def test_data_v4_space_and_sheet_all_case_insensitive(tmp_path: Path) -> None:
    p = tmp_path / "v.xlsx"
    _write_variants_xlsx(p, sheet_title="All", label="Data v4", gap_rows=2)
    report = load_and_analyze(str(p))
    assert report.summary_metrics.get("采样点数") == 5


def _write_with_frameinfo_sheet(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "all"
    ws.append(["Data_v4"])
    ws.append(["Time(ms)", "FPS"])
    for i in range(20):
        ws.append([float(i * 1000), 59.0])
    wf = wb.create_sheet("@FrameInfo")
    wf.append(["Time(ms)", "FrameTime(ms)"])
    for i in range(40):
        ms = 12.0 if i != 10 else 45.0
        wf.append([float(i * 16), ms])
    wb.save(path)


def test_frameinfo_merged_into_report(tmp_path: Path) -> None:
    p = tmp_path / "fi.xlsx"
    _write_with_frameinfo_sheet(p)
    report = load_and_analyze(str(p))
    assert report.frame_stats is not None
    assert report.frame_stats.count == 40
    assert report.frame_stats.max_ms >= 45.0
    assert any("FrameInfo" in f.title for f in report.findings)


def test_compare_reports_package_mismatch_warning(tmp_path: Path) -> None:
    def one(pkg: str) -> AnalysisReport:
        return AnalysisReport(
            session=SessionSummary(package_name=pkg, device_name="d"),
            summary_metrics={"采样点数": 10, "x": 1},
            findings=[
                Finding(
                    id="f1",
                    category=FindingCategory.drop,
                    severity=FindingSeverity.info,
                    title="t",
                    detail="d",
                ),
            ],
            recommendations=[Recommendation(id="r1", finding_ids=[], text="t", category="通用")],
        )

    pair = compare_reports(one("com.a.game"), one("com.b.game"))
    assert pair.warnings
    assert any("包名不一致" in w for w in pair.warnings)


def test_fallback_without_marker_smalljank_headers(tmp_path: Path) -> None:
    """无 Data_v4 字样时仍可通过列名命中表头。"""
    p = tmp_path / "no_label.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["x"])
    ws.append(["Time(ms)", "FPS", "BigJank"])
    for i in range(3):
        ws.append([float(i * 100), 60.0, 0])
    wb.save(p)
    report = load_and_analyze(str(p))
    assert report.summary_metrics.get("采样点数") == 3
