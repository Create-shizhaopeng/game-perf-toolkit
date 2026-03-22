"""安全打开 PerfDog xlsx / xlsm，探测 Data_v4 表头行。"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from toolkit.core.perfdog.column_aliases import map_column_name
from toolkit.core.perfdog.errors import PerfDogParseError, PerfDogUnsupportedError

logger = logging.getLogger(__name__)

# 部分机型导出在 Data_v4 前有大量说明行，适当放大扫描上限
_MAX_SCAN_ROWS = 3000


def safe_load_workbook(path: str, *, read_only: bool = False):
    """打开工作簿，``data_only=True``、不执行宏。

    **重要**：对部分 PerfDog 导出，``read_only=True`` 时 openpyxl 的流式行迭代
    **每行只返回首个有值单元格**（宽表被压成单列），会导致无法识别 Data_v4 表头。
    因此默认使用 **read_only=False**（典型 ``all`` 表行数有限，内存可接受）。
    超大文件如需优化可再引入分块策略。
    """
    p = Path(path)
    if not p.exists():
        raise PerfDogParseError(f"文件不存在: {path}")
    suf = p.suffix.lower()
    if suf not in (".xlsx", ".xlsm"):
        raise PerfDogParseError("仅支持 .xlsx / .xlsm 格式的 PerfDog 导出")

    try:
        return load_workbook(filename=path, read_only=read_only, data_only=True, keep_links=False)
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypted" in msg:
            raise PerfDogUnsupportedError("工作簿已加密或受保护，无法读取") from e
        raise PerfDogParseError(f"无法打开 Excel 文件: {e}") from e


def list_sheet_names(wb) -> list[str]:
    return list(wb.sheetnames)


def _cell_labels_data_v4(cell: object) -> bool:
    """识别「Data_v4 / Data v4 / DATA V4」等标签（单元格可能被合并或带说明文字）。"""
    if cell is None:
        return False
    s = str(cell).strip().lower()
    if not s:
        return False
    # 去掉空白与常见分隔符后是否包含 datav4
    collapsed = re.sub(r"[\s_\-–—]+", "", s)
    if "datav4" in collapsed:
        return True
    # 宽松：data … v4
    if re.search(r"data\s*[_\s\-–—]*\s*v\s*4\b", s):
        return True
    return False


def _row_header_score(values: tuple[Any, ...]) -> int:
    score = 0
    for c in values:
        if c is None:
            continue
        raw = str(c).strip()
        if map_column_name(raw):
            score += 1
    return score


def _row_has_time_and_fps(values: tuple[Any, ...]) -> bool:
    has_t = False
    has_f = False
    for c in values:
        if c is None:
            continue
        m = map_column_name(str(c).strip())
        if m == "time_ms":
            has_t = True
        if m == "fps":
            has_f = True
    return has_t and has_f


def _pick_header_after_marker(
    rows: list[tuple[Any, ...]],
    marker_row_index: int,
    max_lookahead: int = 24,
) -> tuple[int, list[str]] | None:
    """在 Data_v4 标记行之后若干行内，选取最像表头的一行。"""
    n = len(rows)
    start = marker_row_index + 1
    end = min(marker_row_index + 1 + max_lookahead, n)
    best_j: int | None = None
    best_score = 0
    best_headers: list[str] = []

    for j in range(start, end):
        row = rows[j]
        score = _row_header_score(row)
        headers = [str(c).strip() if c is not None else "" for c in row]
        if score > best_score:
            best_score = score
            best_j = j
            best_headers = headers
        if score >= 2 or _row_has_time_and_fps(row):
            return j, headers

    if best_j is not None and (best_score >= 2 or _row_has_time_and_fps(rows[best_j])):
        return best_j, best_headers

    return None


def find_data_v4_header_row(
    sheet,
    max_scan_rows: int = _MAX_SCAN_ROWS,
) -> tuple[int, list[str]] | None:
    """返回 (0-based 表头行索引, 原始表头字符串列表)。"""
    rows: list[tuple[Any, ...]] = []
    for _, row in enumerate(sheet.iter_rows(max_row=max_scan_rows, values_only=True)):
        rows.append(tuple(row))

    # 1) 显式 Data_v4 标记后的表头（允许中间有空行）
    for i, row in enumerate(rows):
        if not any(_cell_labels_data_v4(c) for c in row):
            continue
        picked = _pick_header_after_marker(rows, i)
        if picked:
            return picked

    # 2) 未找到文字标记时：取可映射列命中数最多的行（至少 2 列，或同时含 time+fps）
    best_row: int | None = None
    best_headers: list[str] = []
    best_score = 0
    for i, row in enumerate(rows):
        score = _row_header_score(row)
        better = score > best_score
        if not better and score == best_score and score > 0 and best_row is not None:
            if _row_has_time_and_fps(row) and not _row_has_time_and_fps(rows[best_row]):
                better = True
        if better:
            best_score = score
            best_row = i
            best_headers = [str(c).strip() if c is not None else "" for c in row]

    if best_row is not None and best_headers:
        if best_score >= 2 or _row_has_time_and_fps(rows[best_row]):
            return best_row, best_headers

    return None


def _sheet_probe_order(names: list[str]) -> list[str]:
    """优先扫描名为 all 的工作表（忽略大小写），其余按原顺序。"""
    if not names:
        return []
    lower_map = {n.lower().strip(): n for n in names}
    ordered: list[str] = []
    if "all" in lower_map:
        ordered.append(lower_map["all"])
    for n in names:
        if n not in ordered:
            ordered.append(n)
    return ordered


def probe_workbook(path: str) -> tuple[str, int, list[str]]:
    """返回 (工作表名, 表头行 0-based 索引, 原始表头列表)。"""
    wb = safe_load_workbook(path)
    try:
        names = list_sheet_names(wb)
        if not names:
            raise PerfDogParseError("工作簿不包含任何工作表")

        for sheet_name in _sheet_probe_order(names):
            sheet = wb[sheet_name]
            found = find_data_v4_header_row(sheet)
            if found:
                header_row_idx, headers = found
                logger.debug(
                    "Data_v4 表头定位: sheet=%s row=%s headers=%s",
                    sheet_name,
                    header_row_idx,
                    headers[:8],
                )
                return sheet_name, header_row_idx, headers

        raise PerfDogParseError(
            "无法在导出中找到 Data_v4 数据表头。"
            "若文件中确有 Data_v4：请确认表头行含 Time/FPS 等列（或升级工具版本）；"
            "也可检查 Data_v4 与表头之间是否间隔过多行（>24）。",
        )
    finally:
        wb.close()
