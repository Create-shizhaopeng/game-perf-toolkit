"""@ThreadCpuUsageData 工作表 → DataFrame（read_only 流式）。"""

from __future__ import annotations

import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from toolkit.core.perfdog.column_aliases import map_column_name
from toolkit.core.perfdog.report_types import AnalyzeOptions


def _find_thread_sheet(names: list[str]) -> str | None:
    for n in names:
        low = n.lower().replace("@", "")
        if "threadcpu" in low or "thread_cpu" in low:
            return n
    return None


def _sanitize_thread_col(name: str, idx: int) -> str:
    s = re.sub(r"[^\w\u4e00-\u9fff]+", "_", name.strip())[:64]
    if not s:
        s = f"col_{idx}"
    return f"thread::{s}"


def parse_thread_cpu(path: str, options: AnalyzeOptions) -> pd.DataFrame | None:
    """若存在线程 CPU 表则返回带 ``time_ms`` 与 ``thread::*`` 列的 DataFrame；否则 None。"""
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = _find_thread_sheet(list(wb.sheetnames))
        if not sheet_name:
            return None
        ws = wb[sheet_name]
        header_row: tuple[Any, ...] | None = None
        header_excel_row = 1
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=48, values_only=True),
            start=1,
        ):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            headers = [str(c).strip() if c is not None else "" for c in row]
            mapped = [map_column_name(h) for h in headers if h]
            if "time_ms" in mapped:
                header_row = row
                header_excel_row = i
                break
        if header_row is None:
            return None

        raw_headers = [str(c).strip() if c is not None else "" for c in header_row]
        internal: list[str] = []
        used: set[str] = set()
        for j, h in enumerate(raw_headers):
            if not h:
                internal.append(f"_empty_{j}")
                continue
            m = map_column_name(h)
            if m == "time_ms" and "time_ms" not in used:
                internal.append("time_ms")
                used.add("time_ms")
            elif m and m != "time_ms" and m not in used:
                internal.append(m)
                used.add(m)
            else:
                name = _sanitize_thread_col(h, j)
                while name in used:
                    name = f"{name}_{j}"
                internal.append(name)
                used.add(name)

        rows: list[list[Any]] = []
        count = 0
        for row in ws.iter_rows(min_row=header_excel_row + 1, values_only=True):
            if count >= options.max_frame_rows:
                break
            if not row:
                continue
            pad = list(row) + [None] * max(0, len(internal) - len(row))
            rows.append(pad[: len(internal)])
            count += 1

        if not rows:
            return None
        df = pd.DataFrame(rows, columns=internal)
        if "time_ms" not in df.columns:
            return None
        df["time_ms"] = pd.to_numeric(df["time_ms"], errors="coerce")
        df = df.dropna(subset=["time_ms"])
        return df
    finally:
        wb.close()
