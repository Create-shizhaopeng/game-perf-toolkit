"""@FrameInfo 工作表流式聚合（read_only，大表防 OOM）。"""

from __future__ import annotations

import math
from typing import Any

from openpyxl import load_workbook

from toolkit.core.perfdog.column_aliases import map_column_name
from toolkit.core.perfdog.report_types import AnalyzeOptions, FrameStats


def _find_frameinfo_sheet(names: list[str]) -> str | None:
    for n in names:
        low = n.lower().replace("@", "").strip()
        if "frameinfo" in low:
            return n
    return None


def _header_indices(header_row: tuple[Any, ...]) -> tuple[int | None, int | None]:
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    col_time: int | None = None
    col_ft: int | None = None
    for j, h in enumerate(headers):
        if not h:
            continue
        m = map_column_name(h)
        if m == "time_ms":
            col_time = j
        if m in ("frame_time_ms", "inter_frame"):
            col_ft = j
        low = h.lower().replace(" ", "")
        if col_ft is None and ("frametime" in low or "frameduration" in low):
            col_ft = j
    return col_time, col_ft


def parse_frameinfo(
    path: str,
    options: AnalyzeOptions,
    target_fps: int,
) -> tuple[FrameStats | None, str | None]:
    """扫描 @FrameInfo；返回 (FrameStats, 截断/缺表提示)。"""
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        names = list(wb.sheetnames)
        sheet_name = _find_frameinfo_sheet(names)
        if not sheet_name:
            return None, None

        ws = wb[sheet_name]
        header_row: tuple[Any, ...] | None = None
        header_excel_row = 1
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=32, values_only=True),
            start=1,
        ):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            ct, cf = _header_indices(row)
            if ct is not None and cf is not None:
                header_row = row
                header_excel_row = i
                break
            if cf is not None and ct is None:
                header_row = row
                header_excel_row = i
        if header_row is None:
            return None, "已发现 @FrameInfo 工作表，但未识别到帧耗时列（需 FrameTime / InterFrame 等）。"

        col_time, col_ft = _header_indices(header_row)
        if col_ft is None:
            return None, "@FrameInfo 表头缺少可映射的帧耗时列。"

        if col_time is None:
            col_time = 0

        durations: list[float] = []
        times_at: list[float] = []
        max_rows = options.max_frame_rows
        truncated = False
        count = 0

        for row in ws.iter_rows(min_row=header_excel_row + 1, values_only=True):
            if count >= max_rows:
                truncated = True
                break
            if not row or col_ft >= len(row):
                continue
            raw_ft = row[col_ft]
            if raw_ft is None or (isinstance(raw_ft, float) and math.isnan(raw_ft)):
                continue
            try:
                ft = float(raw_ft)
            except (TypeError, ValueError):
                continue
            if ft < 0 or ft > 1_000_000:
                continue
            t_at = 0.0
            if col_time is not None and col_time < len(row) and row[col_time] is not None:
                try:
                    t_at = float(row[col_time])
                except (TypeError, ValueError):
                    t_at = float(count)
            durations.append(ft)
            times_at.append(t_at)
            count += 1

        if not durations:
            return None, "@FrameInfo 未读到有效帧耗时数据。"

        warn: str | None = None
        if truncated:
            warn = (
                f"@FrameInfo 仅处理前 {max_rows} 行（达到 max_frame_rows 上限），"
                "p99/占比等为截断样本上的估计。"
            )

        arr = sorted(durations)
        n = len(arr)
        mean_ms = sum(arr) / n
        p99_ms = arr[int(math.ceil(0.99 * n)) - 1] if n else 0.0
        max_ms = arr[-1]
        imax = max(range(len(durations)), key=lambda i: durations[i])
        max_frame_at_ms = times_at[imax] if times_at else None

        budget = (1000.0 / max(target_fps, 1)) * 2.0
        over_budget = sum(1 for x in durations if x > budget)

        stats = FrameStats(
            count=n,
            mean_ms=round(mean_ms, 4),
            p99_ms=round(float(p99_ms), 4),
            max_ms=round(float(max_ms), 4),
            over_budget_count=over_budget,
            max_frame_time_ms=round(float(max_ms), 4),
            max_frame_at_ms=max_frame_at_ms,
        )
        return stats, warn
    finally:
        wb.close()
