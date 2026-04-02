"""Perfetto Capture 模块 — 帧率数据导出

将帧率数据导出为 PerfDog 兼容的 xlsx 格式。
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook

from .models import CaptureRegion, FrameExportRow

if TYPE_CHECKING:
    from .fps_chart import FpsChartData


PERFDOG_HEADERS = [
    "num",
    "Time(ms)",
    "Abs_time(ms)",
    "Mono_time(ms)",
    "Label",
    "Notes",
    "FPS",
    "Smooth(%)",
    "1%Low(FPS)",
    "Tiny-Jank(/10min)",
    "Small-Jank(/10min)",
    "Jank(/10min)",
    "BigJank(/10min)",
    "Stutter(%)",
    "Inter_frame(ms)",
]


def export_to_xlsx(
    data: "FpsChartData",
    regions: list[CaptureRegion],
    output_path: Path,
) -> None:
    """导出帧率数据到 xlsx 文件。

    Args:
        data: FPS 图表数据
        regions: 抓取选区列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_v4"

    for col, header in enumerate(PERFDOG_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    timestamps, fps_values = data.get_curve_data()
    if len(timestamps) == 0:
        wb.save(output_path)
        return

    start_time_sec = timestamps[0] if len(timestamps) > 0 else 0

    for idx, (ts, fps) in enumerate(zip(timestamps, fps_values)):
        row_idx = idx + 2
        time_ms = int((ts - start_time_sec) * 1000)

        label = _get_region_label_at(regions, ts, start_time_sec)

        row = FrameExportRow(
            num=idx + 1,
            time=time_ms,
            label=label,
            fps=round(fps, 1) if label else None,
        )

        ws.cell(row=row_idx, column=1, value=row.num)
        ws.cell(row=row_idx, column=2, value=row.time)
        ws.cell(row=row_idx, column=3, value=row.abs_time)
        ws.cell(row=row_idx, column=4, value=row.mono_time)
        ws.cell(row=row_idx, column=5, value=row.label)
        ws.cell(row=row_idx, column=6, value=row.notes)
        ws.cell(row=row_idx, column=7, value=row.fps)
        ws.cell(row=row_idx, column=8, value=row.smooth)
        ws.cell(row=row_idx, column=9, value=row.low_1_percent_fps)
        ws.cell(row=row_idx, column=10, value=row.tiny_jank)
        ws.cell(row=row_idx, column=11, value=row.small_jank)
        ws.cell(row=row_idx, column=12, value=row.jank)
        ws.cell(row=row_idx, column=13, value=row.big_jank)
        ws.cell(row=row_idx, column=14, value=row.stutter_percent)
        ws.cell(row=row_idx, column=15, value=row.inter_frame)

    wb.save(output_path)


def _get_region_label_at(
    regions: list[CaptureRegion],
    timestamp_sec: float,
    start_time_sec: float,
) -> str:
    """获取指定时间点的选区标签。

    Args:
        regions: 选区列表
        timestamp_sec: 相对时间戳（秒）
        start_time_sec: 数据起始时间戳

    Returns:
        选区标签，非抓取选区返回空字符串
    """
    if not regions:
        return ""

    elapsed_sec = timestamp_sec - start_time_sec
    first_region_start = regions[0].start_time

    for region in regions:
        region_start_sec = (region.start_time - first_region_start).total_seconds()
        if region.end_time:
            region_end_sec = (region.end_time - first_region_start).total_seconds()
        else:
            region_end_sec = float("inf")

        if region_start_sec <= elapsed_sec <= region_end_sec:
            return region.label if region.is_capture else ""

    return ""
