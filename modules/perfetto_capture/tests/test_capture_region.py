"""选区与导出测试"""

import datetime
import tempfile
from pathlib import Path

import pytest

from modules.perfetto_capture.src.models import CaptureRegion, FrameExportRow
from modules.perfetto_capture.src.frame_exporter import export_to_xlsx, _get_region_label_at


class MockFpsChartData:
    """模拟 FPS 图表数据。"""

    def __init__(self, timestamps: list[float], fps_values: list[float]):
        self._timestamps = timestamps
        self._fps_values = fps_values

    def get_curve_data(self):
        import numpy as np
        return np.array(self._timestamps), np.array(self._fps_values)


class TestCaptureRegion:
    """测试 CaptureRegion 模型。"""

    def test_create_capture_region(self):
        """测试创建抓取选区。"""
        now = datetime.datetime.now()
        region = CaptureRegion(
            start_time=now,
            is_capture=True,
            label="capture1",
        )

        assert region.start_time == now
        assert region.end_time is None
        assert region.is_capture is True
        assert region.label == "capture1"

    def test_create_non_capture_region(self):
        """测试创建非抓取选区。"""
        now = datetime.datetime.now()
        region = CaptureRegion(
            start_time=now,
            is_capture=False,
            label="",
        )

        assert region.is_capture is False
        assert region.label == ""

    def test_region_with_end_time(self):
        """测试带结束时间的选区。"""
        start = datetime.datetime.now()
        end = start + datetime.timedelta(seconds=10)

        region = CaptureRegion(
            start_time=start,
            end_time=end,
            is_capture=True,
            label="capture1",
        )

        assert region.end_time == end
        duration = (region.end_time - region.start_time).total_seconds()
        assert duration == 10


class TestFrameExportRow:
    """测试 FrameExportRow 模型。"""

    def test_create_export_row(self):
        """测试创建导出行。"""
        row = FrameExportRow(
            num=1,
            time=1000,
            label="capture1",
            fps=60.5,
        )

        assert row.num == 1
        assert row.time == 1000
        assert row.label == "capture1"
        assert row.fps == 60.5
        assert row.abs_time is None
        assert row.smooth is None

    def test_export_row_non_capture(self):
        """测试非抓取选区行。"""
        row = FrameExportRow(
            num=1,
            time=1000,
            label="",
            fps=None,
        )

        assert row.label == ""
        assert row.fps is None


class TestGetRegionLabel:
    """测试选区标签获取。"""

    def test_no_regions(self):
        """测试无选区。"""
        label = _get_region_label_at([], 10.0, 0.0)
        assert label == ""

    def test_in_capture_region(self):
        """测试在抓取选区内。"""
        start = datetime.datetime.now()
        regions = [
            CaptureRegion(
                start_time=start,
                end_time=start + datetime.timedelta(seconds=30),
                is_capture=True,
                label="capture1",
            ),
        ]

        label = _get_region_label_at(regions, 15.0, 0.0)
        assert label == "capture1"

    def test_in_non_capture_region(self):
        """测试在非抓取选区内。"""
        start = datetime.datetime.now()
        regions = [
            CaptureRegion(
                start_time=start,
                end_time=start + datetime.timedelta(seconds=30),
                is_capture=False,
                label="",
            ),
        ]

        label = _get_region_label_at(regions, 15.0, 0.0)
        assert label == ""

    def test_multiple_regions(self):
        """测试多个选区。"""
        start = datetime.datetime.now()
        regions = [
            CaptureRegion(
                start_time=start,
                end_time=start + datetime.timedelta(seconds=10),
                is_capture=True,
                label="capture1",
            ),
            CaptureRegion(
                start_time=start + datetime.timedelta(seconds=10),
                end_time=start + datetime.timedelta(seconds=20),
                is_capture=False,
                label="",
            ),
            CaptureRegion(
                start_time=start + datetime.timedelta(seconds=20),
                end_time=start + datetime.timedelta(seconds=40),
                is_capture=True,
                label="capture2",
            ),
        ]

        assert _get_region_label_at(regions, 5.0, 0.0) == "capture1"
        assert _get_region_label_at(regions, 15.0, 0.0) == ""
        assert _get_region_label_at(regions, 30.0, 0.0) == "capture2"


class TestExportToXlsx:
    """测试 xlsx 导出。"""

    def test_export_empty_data(self):
        """测试导出空数据。"""
        data = MockFpsChartData([], [])
        regions = []

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = Path(f.name)

        try:
            export_to_xlsx(data, regions, output_path)
            assert output_path.exists()
        finally:
            output_path.unlink()

    def test_export_with_data(self):
        """测试导出有数据。"""
        timestamps = [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]
        fps_values = [60.0, 59.0, 58.0, 60.0, 61.0, 60.0]
        data = MockFpsChartData(timestamps, fps_values)

        start = datetime.datetime.now()
        regions = [
            CaptureRegion(
                start_time=start,
                end_time=start + datetime.timedelta(seconds=10),
                is_capture=True,
                label="capture1",
            ),
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = Path(f.name)

        try:
            export_to_xlsx(data, regions, output_path)
            assert output_path.exists()
            assert output_path.stat().st_size > 0

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            assert ws.title == "Data_v4"
            assert ws.cell(row=1, column=1).value == "num"
            assert ws.cell(row=1, column=7).value == "FPS"

            assert ws.cell(row=2, column=1).value == 1
        finally:
            output_path.unlink()

    def test_export_file_format(self):
        """测试导出文件格式。"""
        timestamps = [0.0, 0.5, 1.0]
        fps_values = [60.0, 55.0, 58.0]
        data = MockFpsChartData(timestamps, fps_values)
        regions = []

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = Path(f.name)

        try:
            export_to_xlsx(data, regions, output_path)

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            headers = [ws.cell(row=1, column=i).value for i in range(1, 16)]
            expected = [
                "num", "Time(ms)", "Abs_time(ms)", "Mono_time(ms)",
                "Label", "Notes", "FPS", "Smooth(%)", "1%Low(FPS)",
                "Tiny-Jank(/10min)", "Small-Jank(/10min)", "Jank(/10min)",
                "BigJank(/10min)", "Stutter(%)", "Inter_frame(ms)",
            ]
            assert headers == expected
        finally:
            output_path.unlink()
